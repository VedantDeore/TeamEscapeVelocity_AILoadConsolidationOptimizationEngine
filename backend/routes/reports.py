"""Report generation routes — PDF / CSV / Excel export."""

from flask import Blueprint, jsonify, Response, request
from models.supabase_client import get_supabase
import csv
import io
import json
from datetime import datetime

reports_bp = Blueprint("reports", __name__)


@reports_bp.route("/api/reports", methods=["GET"])
def list_reports():
    sb     = get_supabase()
    result = sb.table("reports").select("*").execute()
    return jsonify(result.data)


VALID_REPORT_TYPES = {"consolidation", "utilization", "cost", "carbon", "route"}

_REPORT_META = {
    "consolidation": {
        "title": "Daily Consolidation Summary",
        "headers": ["cluster_id", "vehicle_name", "shipment_count",
                     "total_weight_kg", "utilization_pct",
                     "route_distance_km", "estimated_cost_inr", "estimated_co2_kg", "status"],
        "filename": "lorri_consolidation_report",
    },
    "utilization": {
        "title": "Vehicle Utilization Report",
        "headers": ["vehicle_name", "utilization_pct", "total_weight_kg",
                     "total_volume_m3", "route_distance_km", "estimated_cost_inr"],
        "filename": "lorri_utilization_report",
    },
    "cost": {
        "title": "Cost Savings Analysis",
        "headers": ["plan_name", "trips_before", "trips_after", "trips_saved",
                     "cost_before_inr", "cost_after_inr", "cost_saved_inr",
                     "co2_before_kg", "co2_after_kg", "co2_saved_kg"],
        "filename": "lorri_cost_savings_report",
    },
    "carbon": {
        "title": "Carbon Impact Report",
        "headers": ["month", "co2_before_kg", "co2_after_kg", "savings_kg"],
        "filename": "lorri_carbon_report",
    },
    "route": {
        "title": "Route Efficiency Report",
        "headers": ["route_id", "vehicle_name", "total_stops", "distance_km",
                     "estimated_time_min", "fuel_cost_inr", "co2_kg"],
        "filename": "lorri_route_efficiency_report",
    },
}

@reports_bp.route("/api/reports/<report_type>", methods=["GET"])
def get_report(report_type):
    if report_type not in VALID_REPORT_TYPES:
        return jsonify({"error": f"Unknown report type: {report_type}", "valid_types": sorted(VALID_REPORT_TYPES)}), 400
    sb     = get_supabase()
    result = sb.table("reports").select("*").eq("type", report_type).execute()
    return jsonify(result.data)


@reports_bp.route("/api/reports/<report_type>/download", methods=["GET"])
def download_report(report_type):
    """Generate and stream a PDF, CSV, or Excel report."""
    if report_type not in VALID_REPORT_TYPES:
        return jsonify({"error": f"Unknown report type: {report_type}"}), 400

    sb  = get_supabase()
    fmt = request.args.get("format", "csv").lower()

    builders = {
        "consolidation": _build_consolidation_report,
        "utilization":   _build_utilization_report,
        "cost":          _build_cost_report,
        "carbon":        _build_carbon_report,
        "route":         _build_route_report,
    }
    rows = builders[report_type](sb)
    meta = _REPORT_META[report_type]
    headers_list = meta["headers"]
    base_name    = meta["filename"]
    title        = meta["title"]

    if fmt == "pdf":
        return _respond_pdf(rows, headers_list, base_name, title)
    elif fmt in ("excel", "xlsx"):
        return _respond_excel(rows, headers_list, base_name, title)
    elif fmt == "json":
        return Response(
            json.dumps(rows, indent=2),
            mimetype="application/json",
            headers={"Content-Disposition": f"attachment; filename={base_name}.json"},
        )
    else:
        return _respond_csv(rows, headers_list, base_name)


def _respond_csv(rows, headers_list, base_name):
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=headers_list, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={base_name}.csv"},
    )


def _respond_pdf(rows, headers_list, base_name, title):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buf = io.BytesIO()
    page = landscape(A4) if len(headers_list) > 6 else A4
    doc = SimpleDocTemplate(buf, pagesize=page,
                            topMargin=40, bottomMargin=30,
                            leftMargin=30, rightMargin=30)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"<b>{title}</b>", styles["Title"]))
    elements.append(Paragraph(
        f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')} &nbsp;|&nbsp; "
        f"Lorri AI Load Consolidation Engine",
        styles["Normal"],
    ))
    elements.append(Spacer(1, 16))

    display_headers = [h.replace("_", " ").title() for h in headers_list]
    table_data = [display_headers]
    for row in rows:
        table_data.append([str(row.get(h, "")) for h in headers_list])

    t = Table(table_data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0a2540")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, -1), 7.5),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e3e8ee")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fc")]),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(
        f"<i>Total rows: {len(rows)} &nbsp;|&nbsp; Confidential — Lorri Logistics AI</i>",
        styles["Normal"],
    ))

    doc.build(elements)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={base_name}.pdf"},
    )


def _respond_excel(rows, headers_list, base_name, title):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="0A2540", end_color="0A2540", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="E3E8EE"),
        right=Side(style="thin", color="E3E8EE"),
        top=Side(style="thin", color="E3E8EE"),
        bottom=Side(style="thin", color="E3E8EE"),
    )
    alt_fill = PatternFill(start_color="F8F9FC", end_color="F8F9FC", fill_type="solid")

    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers_list))
    ws["A1"].font = Font(bold=True, size=14, color="0A2540")
    ws.append([f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}"])
    ws.append([])

    display_headers = [h.replace("_", " ").title() for h in headers_list]
    ws.append(display_headers)
    for col_idx, _ in enumerate(display_headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for r_idx, row in enumerate(rows, 5):
        vals = [row.get(h, "") for h in headers_list]
        ws.append(vals)
        for c_idx in range(1, len(headers_list) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            if r_idx % 2 == 1:
                cell.fill = alt_fill

    for col_idx, h in enumerate(headers_list, 1):
        max_len = max(len(str(h)), *(len(str(row.get(h, ""))) for row in rows)) if rows else len(h)
        ws.column_dimensions[ws.cell(row=4, column=col_idx).column_letter].width = min(max_len + 4, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={base_name}.xlsx"},
    )


# ── report builders ────────────────────────────────────────

def _build_consolidation_report(sb) -> list:
    clusters = sb.table("clusters").select("*").execute().data or []
    all_cs = sb.table("cluster_shipments").select("cluster_id").execute().data or []
    counts: dict = {}
    for cs in all_cs:
        cid = cs["cluster_id"]
        counts[cid] = counts.get(cid, 0) + 1
    return [{
        "cluster_id":        c["id"][:8],
        "vehicle_name":      c.get("vehicle_name", ""),
        "shipment_count":    counts.get(c["id"], 0),
        "total_weight_kg":   c.get("total_weight", 0),
        "utilization_pct":   c.get("utilization_pct", 0),
        "route_distance_km": c.get("route_distance_km", 0),
        "estimated_cost_inr": c.get("estimated_cost", 0),
        "estimated_co2_kg":  c.get("estimated_co2", 0),
        "status":            c.get("status", "pending"),
    } for c in clusters]


def _build_utilization_report(sb) -> list:
    clusters = sb.table("clusters").select("*").execute().data or []
    return [{
        "vehicle_name":      c.get("vehicle_name", ""),
        "utilization_pct":   c.get("utilization_pct", 0),
        "total_weight_kg":   c.get("total_weight", 0),
        "total_volume_m3":   c.get("total_volume", 0),
        "route_distance_km": c.get("route_distance_km", 0),
        "estimated_cost_inr": c.get("estimated_cost", 0),
    } for c in clusters]


def _build_cost_report(sb) -> list:
    plans = sb.table("consolidation_plans").select("*").order("created_at", desc=True).execute().data or []
    return [{
        "plan_name":       p.get("name", ""),
        "trips_before":    p.get("trips_before", 0),
        "trips_after":     p.get("trips_after", 0),
        "trips_saved":     (p.get("trips_before", 0) or 0) - (p.get("trips_after", 0) or 0),
        "cost_before_inr": p.get("total_cost_before", 0),
        "cost_after_inr":  p.get("total_cost_after", 0),
        "cost_saved_inr":  (p.get("total_cost_before", 0) or 0) - (p.get("total_cost_after", 0) or 0),
        "co2_before_kg":   p.get("co2_before", 0),
        "co2_after_kg":    p.get("co2_after", 0),
        "co2_saved_kg":    (p.get("co2_before", 0) or 0) - (p.get("co2_after", 0) or 0),
    } for p in plans]


def _build_carbon_report(sb) -> list:
    monthly = sb.table("carbon_monthly").select("*").execute().data or []
    return [{
        "month":         m.get("month", ""),
        "co2_before_kg": m.get("co2_before", 0),
        "co2_after_kg":  m.get("co2_after", 0),
        "savings_kg":    m.get("savings", 0),
    } for m in monthly]


def _build_route_report(sb) -> list:
    routes = sb.table("routes").select("*").execute().data or []
    return [{
        "route_id":          r.get("id", "")[:8],
        "vehicle_name":      r.get("vehicle_name", ""),
        "total_stops":       len(r.get("points", [])) if isinstance(r.get("points"), list) else 0,
        "distance_km":       r.get("total_distance_km", 0),
        "estimated_time_min": r.get("estimated_time_min", 0),
        "fuel_cost_inr":     r.get("fuel_cost", 0),
        "co2_kg":            r.get("co2_kg", 0),
    } for r in routes]
